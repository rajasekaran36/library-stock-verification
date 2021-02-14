from openpyxl import load_workbook, workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
fields = ['Access No','Title','Volume','Author 1','Author 2','Author 3','ISBN','Department','Edition','Publisher','Supplier','Cover Type','Issue Type','Pages','Published Year','Shelf No','Row No','Source','Purchase DateInvoice No','Key Words','Book Location','Sur Name','Amount']

def rec_line_to_dictionary(record_line):
    record_dict_format = {}.fromkeys(fields)
    datas = record_line
    for i in range(len(datas)):
        record_dict_format[fields[i]]=datas[i]
    return record_dict_format


def rec_lines_to_dictionary(list_records):
    all_records_dict = {}
    for record in list_records:
        all_records_dict[record[0]] = rec_line_to_dictionary(record)
    return all_records_dict

def file_of_records_to_list(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    list_of_records=[]

    for row in sheet.rows:
        lis=[]
        for cell in row:
            lis.append(cell.value)
        list_of_records.append(tuple(lis))
    return list_of_records

def available_access_numbers(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    list_of_keys_available=[]
    for row in sheet.rows:
        for cell in row:
            list_of_keys_available.append(cell.value)
    return list_of_keys_available


full_stock_record_list = file_of_records_to_list('full-stock.xlsx')
full_stock_dictionary = rec_lines_to_dictionary(full_stock_record_list)
print(len(full_stock_dictionary))
all_keys = full_stock_dictionary.keys()
avl_keys = available_access_numbers('aval-stock.xlsx')
missing_keys = [key for key in all_keys if key not in avl_keys]

print('total stock:',len(all_keys))
print('aval books:',len(avl_keys))
print('missing books:',len(missing_keys))

workbook = Workbook()
worksheet = workbook.active
for missedkey in missing_keys:
    record = full_stock_dictionary.get(missedkey)
    #print(list(record.values()))
    worksheet.append(list(record.values()))
workbook.save(filename ='missed-stock.xlsx')

print('verification done')